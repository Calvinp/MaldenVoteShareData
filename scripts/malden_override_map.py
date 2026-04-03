from __future__ import annotations

import base64
import io
import json
import math
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageEnhance, ImageFont, ImageOps
from shapely.geometry import MultiPolygon, Polygon, shape
from shapely.ops import unary_union


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "RawData" / "malden_override_results_verified.xlsx"
PRECINCTS_PATH = ROOT / "RawData" / "malden_subprecincts_official.geojson"
OUTPUT_DIR = ROOT / "Output"
TILE_CACHE_DIR = ROOT / ".cache" / "tiles"

OSM_TILE_TEMPLATE = "https://tile.openstreetmap.org/{z}/{x}/{y}.png"
TILE_USER_AGENT = "Prop2.5OverrideData/1.0 (side-project local map generator)"

COLOR_STOPS = [
    (0.25, (202, 0, 32)),
    (0.50, (247, 247, 247)),
    (0.75, (5, 113, 176)),
]

DIFF_COLOR_STOPS = [
    (-0.10, (59, 130, 246)),
    (0.00, (245, 245, 235)),
    (0.10, (217, 119, 6)),
]


@dataclass(frozen=True)
class PrecinctResult:
    precinct: str
    q1a_yes_pct: float
    q1b_yes_pct: float
    q1a_yes: int
    q1a_no: int
    q1b_yes: int
    q1b_no: int

    @property
    def ward(self) -> str:
        return self.precinct.split("-")[0]

    @property
    def q1a_minus_q1b_yes_pct(self) -> float:
        return self.q1a_yes_pct - self.q1b_yes_pct


@dataclass(frozen=True)
class Basemap:
    image: Image.Image
    city_mask: Image.Image
    min_world_x: float
    min_world_y: float
    width: int
    height: int
    zoom: int

    def project(self, lon: float, lat: float) -> tuple[float, float]:
        world_x, world_y = lonlat_to_world_pixels(lon, lat, self.zoom)
        return (world_x - self.min_world_x, world_y - self.min_world_y)


@dataclass(frozen=True)
class LegendSpec:
    title: str
    items: list[tuple[str, tuple[int, int, int]]]
    position: str = "top-right"


def normalize_precinct_name(name: str) -> str:
    return name.strip().upper()


def load_precinct_results(workbook_path: Path = WORKBOOK_PATH) -> dict[str, PrecinctResult]:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["By Precinct"]
    results: dict[str, PrecinctResult] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        precinct = row[0]
        if not precinct or normalize_precinct_name(str(precinct)) == "TOTAL":
            continue

        precinct_name = normalize_precinct_name(str(precinct))
        q1a_yes = int(row[2])
        q1a_no = int(row[3])
        q1b_yes = int(row[6])
        q1b_no = int(row[7])

        results[precinct_name] = PrecinctResult(
            precinct=precinct_name,
            q1a_yes_pct=q1a_yes / (q1a_yes + q1a_no),
            q1b_yes_pct=q1b_yes / (q1b_yes + q1b_no),
            q1a_yes=q1a_yes,
            q1a_no=q1a_no,
            q1b_yes=q1b_yes,
            q1b_no=q1b_no,
        )

    return results


def load_precinct_geometries(precincts_path: Path = PRECINCTS_PATH) -> dict[str, Polygon | MultiPolygon]:
    data = json.loads(precincts_path.read_text(encoding="utf-8"))
    geometries: dict[str, Polygon | MultiPolygon] = {}
    for feature in data["features"]:
        name = normalize_precinct_name(feature["properties"]["DIST_NAME"])
        geometries[name] = shape(feature["geometry"])
    return geometries


def validate_join(
    results: dict[str, PrecinctResult], geometries: dict[str, Polygon | MultiPolygon]
) -> tuple[list[str], list[str]]:
    missing_geometries = sorted(set(results) - set(geometries))
    extra_geometries = sorted(set(geometries) - set(results))
    return missing_geometries, extra_geometries


def interpolate_color(percent_yes: float) -> tuple[int, int, int]:
    value = max(0.25, min(0.75, percent_yes))
    for (left_value, left_color), (right_value, right_color) in zip(COLOR_STOPS, COLOR_STOPS[1:]):
        if left_value <= value <= right_value:
            fraction = (value - left_value) / (right_value - left_value)
            return tuple(
                round(left_channel + (right_channel - left_channel) * fraction)
                for left_channel, right_channel in zip(left_color, right_color)
            )
    return COLOR_STOPS[-1][1]


def interpolate_difference_color(diff_value: float) -> tuple[int, int, int]:
    value = max(DIFF_COLOR_STOPS[0][0], min(DIFF_COLOR_STOPS[-1][0], diff_value))
    for (left_value, left_color), (right_value, right_color) in zip(
        DIFF_COLOR_STOPS, DIFF_COLOR_STOPS[1:]
    ):
        if left_value <= value <= right_value:
            fraction = (value - left_value) / (right_value - left_value)
            return tuple(
                round(left_channel + (right_channel - left_channel) * fraction)
                for left_channel, right_channel in zip(left_color, right_color)
            )
    return DIFF_COLOR_STOPS[-1][1]


def interpolate_between_stops(
    value: float, stops: list[tuple[float, tuple[int, int, int]]]
) -> tuple[int, int, int]:
    if value <= stops[0][0]:
        return stops[0][1]
    if value >= stops[-1][0]:
        return stops[-1][1]

    for (left_value, left_color), (right_value, right_color) in zip(stops, stops[1:]):
        if left_value <= value <= right_value:
            fraction = (value - left_value) / (right_value - left_value)
            return tuple(
                round(left_channel + (right_channel - left_channel) * fraction)
                for left_channel, right_channel in zip(left_color, right_color)
            )
    return stops[-1][1]


def build_difference_colorizer(
    results: dict[str, PrecinctResult],
) -> tuple[callable, LegendSpec]:
    values = [result.q1a_minus_q1b_yes_pct for result in results.values()]
    min_value = min(values)
    max_value = max(values)

    if min_value >= 0:
        stops = [
            (min_value, (245, 245, 235)),
            ((min_value + max_value) / 2, (235, 173, 105)),
            (max_value, (217, 119, 6)),
        ]
    elif max_value <= 0:
        stops = [
            (min_value, (59, 130, 246)),
            ((min_value + max_value) / 2, (152, 190, 247)),
            (max_value, (245, 245, 235)),
        ]
    else:
        stops = [
            (min_value, (59, 130, 246)),
            (0.0, (245, 245, 235)),
            (max_value, (230, 120, 55)),
        ]

    def color_fn(value: float) -> tuple[int, int, int]:
        return interpolate_between_stops(value, stops)

    legend_values = sorted({min_value, min_value + (max_value - min_value) * 0.33, min_value + (max_value - min_value) * 0.66, max_value})
    legend = LegendSpec(
        title="Q1A - Q1B yes %",
        items=[(format_pct_point(v), color_fn(v)) for v in legend_values],
        position="bottom-right",
    )
    return color_fn, legend


def build_vote_share_legend(question_label: str) -> LegendSpec:
    levels = [0.25, 0.50, 0.75]
    return LegendSpec(
        title=f"{question_label} yes vote share",
        items=[(f"{round(level * 100):.0f}%", interpolate_color(level)) for level in levels],
        position="bottom-right",
    )


def lonlat_to_world_pixels(lon: float, lat: float, zoom: int) -> tuple[float, float]:
    scale = 256 * (2**zoom)
    x = (lon + 180.0) / 360.0 * scale
    lat_rad = math.radians(lat)
    y = (
        (1.0 - math.log(math.tan(lat_rad) + (1.0 / math.cos(lat_rad))) / math.pi)
        / 2.0
        * scale
    )
    return x, y


def world_pixels_to_tile_bounds(
    min_x: float, min_y: float, max_x: float, max_y: float
) -> tuple[int, int, int, int]:
    return (
        int(math.floor(min_x / 256)),
        int(math.floor(min_y / 256)),
        int(math.floor(max_x / 256)),
        int(math.floor(max_y / 256)),
    )


def fetch_tile(z: int, x: int, y: int, cache_dir: Path = TILE_CACHE_DIR) -> Image.Image:
    cache_dir.mkdir(parents=True, exist_ok=True)
    tile_path = cache_dir / f"{z}_{x}_{y}.png"
    if not tile_path.exists():
        request = urllib.request.Request(
            OSM_TILE_TEMPLATE.format(z=z, x=x, y=y),
            headers={"User-Agent": TILE_USER_AGENT},
        )
        with urllib.request.urlopen(request, timeout=30) as response:
            tile_path.write_bytes(response.read())
    return Image.open(tile_path).convert("RGB")


def build_basemap(
    geometries: dict[str, Polygon | MultiPolygon],
    zoom: int = 15,
    padding: int = 96,
) -> Basemap:
    union = unary_union(list(geometries.values()))
    min_lon, min_lat, max_lon, max_lat = union.bounds
    min_world_x, max_world_y = lonlat_to_world_pixels(min_lon, min_lat, zoom)
    max_world_x, min_world_y = lonlat_to_world_pixels(max_lon, max_lat, zoom)

    padded_min_world_x = min_world_x - padding
    padded_min_world_y = min_world_y - padding
    padded_max_world_x = max_world_x + padding
    padded_max_world_y = max_world_y + padding

    tile_min_x, tile_min_y, tile_max_x, tile_max_y = world_pixels_to_tile_bounds(
        padded_min_world_x, padded_min_world_y, padded_max_world_x, padded_max_world_y
    )

    stitched = Image.new(
        "RGB",
        ((tile_max_x - tile_min_x + 1) * 256, (tile_max_y - tile_min_y + 1) * 256),
        "white",
    )
    for tile_y in range(tile_min_y, tile_max_y + 1):
        for tile_x in range(tile_min_x, tile_max_x + 1):
            tile = fetch_tile(zoom, tile_x, tile_y)
            stitched.paste(tile, ((tile_x - tile_min_x) * 256, (tile_y - tile_min_y) * 256))

    crop_left = round(padded_min_world_x - tile_min_x * 256)
    crop_top = round(padded_min_world_y - tile_min_y * 256)
    crop_right = round(padded_max_world_x - tile_min_x * 256)
    crop_bottom = round(padded_max_world_y - tile_min_y * 256)
    cropped = stitched.crop((crop_left, crop_top, crop_right, crop_bottom))

    grayscale = ImageOps.grayscale(cropped).convert("RGB")
    grayscale = ImageEnhance.Contrast(grayscale).enhance(1.25)
    grayscale = ImageEnhance.Brightness(grayscale).enhance(1.15)

    city_mask = Image.new("L", grayscale.size, 0)
    mask_draw = ImageDraw.Draw(city_mask)
    draw_polygon_geometry(mask_draw, union, lambda lon, lat: (lonlat_to_world_pixels(lon, lat, zoom)[0] - padded_min_world_x, lonlat_to_world_pixels(lon, lat, zoom)[1] - padded_min_world_y), fill=255)

    clipped = Image.new("RGB", grayscale.size, "white")
    clipped.paste(grayscale, mask=city_mask)

    return Basemap(
        image=clipped,
        city_mask=city_mask,
        min_world_x=padded_min_world_x,
        min_world_y=padded_min_world_y,
        width=clipped.width,
        height=clipped.height,
        zoom=zoom,
    )


def draw_polygon_geometry(
    draw: ImageDraw.ImageDraw,
    geometry: Polygon | MultiPolygon,
    projector,
    *,
    fill=None,
    outline=None,
    width: int = 1,
) -> None:
    polygons = list(geometry.geoms) if isinstance(geometry, MultiPolygon) else [geometry]
    for polygon in polygons:
        exterior = [projector(lon, lat) for lon, lat in polygon.exterior.coords]
        draw.polygon(exterior, fill=fill, outline=outline, width=width)


def geometry_to_svg_path(geometry: Polygon | MultiPolygon, projector) -> str:
    polygons = list(geometry.geoms) if isinstance(geometry, MultiPolygon) else [geometry]
    segments: list[str] = []
    for polygon in polygons:
        for ring in [polygon.exterior, *polygon.interiors]:
            coords = [projector(lon, lat) for lon, lat in ring.coords]
            if not coords:
                continue
            first_x, first_y = coords[0]
            parts = [f"M {first_x:.2f} {first_y:.2f}"]
            for x, y in coords[1:]:
                parts.append(f"L {x:.2f} {y:.2f}")
            parts.append("Z")
            segments.append(" ".join(parts))
    return " ".join(segments)


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf"),
        Path("C:/Windows/Fonts/Arialbd.ttf"),
        Path("C:/Windows/Fonts/segoeuib.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def format_pct_point(value: float) -> str:
    return f"{value * 100:+.1f} pts"


def draw_legend(draw: ImageDraw.ImageDraw, image_size: tuple[int, int], legend: LegendSpec) -> None:
    width, height = image_size
    title_font = load_font(max(20, round(width * 0.014)))
    item_font = load_font(max(16, round(width * 0.0105)))
    padding = 18
    swatch = 28
    line_gap = 14
    item_gap = 10

    def text_size(font, text: str) -> tuple[int, int]:
        bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]

    title_width, title_height = text_size(title_font, legend.title)
    item_width = 0
    item_heights = []
    for label, _ in legend.items:
        w, h = text_size(item_font, label)
        item_width = max(item_width, w)
        item_heights.append(h)

    content_width = max(title_width, swatch + line_gap + item_width)
    box_width = padding * 2 + content_width
    box_height = (
        padding * 2
        + title_height
        + item_gap
        + sum(max(swatch, h) for h in item_heights)
        + line_gap * max(0, len(legend.items) - 1)
    )
    left = width - box_width - 28
    top = 28 if legend.position != "bottom-right" else height - box_height - 28

    draw.rounded_rectangle(
        (left, top, left + box_width, top + box_height),
        radius=16,
        fill=(255, 255, 255, 230),
        outline=(90, 90, 90, 220),
        width=2,
    )
    draw.text((left + padding, top + padding), legend.title, fill=(45, 45, 45), font=title_font)

    cursor_y = top + padding + title_height + item_gap
    for (label, color), item_height in zip(legend.items, item_heights):
        swatch_y = cursor_y + max(0, (max(swatch, item_height) - swatch) / 2)
        text_y = cursor_y + max(0, (max(swatch, item_height) - item_height) / 2)
        draw.rounded_rectangle(
            (left + padding, swatch_y, left + padding + swatch, swatch_y + swatch),
            radius=6,
            fill=(*color, 255),
            outline=(120, 120, 120, 200),
            width=1,
        )
        draw.text(
            (left + padding + swatch + line_gap, text_y),
            label,
            fill=(55, 55, 55),
            font=item_font,
        )
        cursor_y += max(swatch, item_height) + line_gap


def draw_title(draw: ImageDraw.ImageDraw, image_size: tuple[int, int], title: str) -> None:
    width, _ = image_size
    title_font = load_font(max(26, round(width * 0.018)))
    bbox = draw.textbbox((0, 0), title, font=title_font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    x = (width - text_width) / 2
    y = 26
    draw.text((x, y), title, fill=(38, 38, 38), font=title_font)
    draw.line((30, y + text_height + 18, width - 30, y + text_height + 18), fill=(185, 185, 185), width=2)


def estimate_svg_legend_box(legend: LegendSpec) -> tuple[int, int]:
    title_width = max(180, len(legend.title) * 12)
    item_width = max((len(label) * 10 for label, _ in legend.items), default=120)
    box_width = max(title_width, 28 + 14 + item_width) + 36
    box_height = 92 + len(legend.items) * 42
    return box_width, box_height


def render_map(
    title: str,
    value_getter,
    results: dict[str, PrecinctResult],
    geometries: dict[str, Polygon | MultiPolygon],
    basemap: Basemap,
    output_stem: Path,
    color_fn,
    legend: LegendSpec | None = None,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    projected = lambda lon, lat: basemap.project(lon, lat)

    map_image = basemap.image.convert("RGBA")
    overlay = Image.new("RGBA", map_image.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)

    for precinct_name, geometry in geometries.items():
        result = results[precinct_name]
        value = value_getter(result)
        fill_rgb = color_fn(value)
        fill_rgba = (*fill_rgb, 190)
        draw_polygon_geometry(
            draw,
            geometry,
            projected,
            fill=fill_rgba,
            outline=(255, 255, 255, 160),
            width=2,
        )

    map_image = Image.alpha_composite(map_image, overlay)

    ward_overlay = Image.new("RGBA", map_image.size, (255, 255, 255, 0))
    ward_draw = ImageDraw.Draw(ward_overlay)
    for ward in sorted({result.ward for result in results.values()}):
        ward_geometry = unary_union(
            [geometries[name] for name, result in results.items() if result.ward == ward]
        )
        draw_polygon_geometry(
            ward_draw,
            ward_geometry,
            projected,
            outline=(70, 70, 70, 220),
            width=5,
        )
    city_geometry = unary_union(list(geometries.values()))
    draw_polygon_geometry(
        ward_draw,
        city_geometry,
        projected,
        outline=(40, 40, 40, 255),
        width=7,
    )
    map_image = Image.alpha_composite(map_image, ward_overlay)

    label_font = load_font(max(18, round(map_image.width * 0.0115)))
    label_draw = ImageDraw.Draw(map_image)
    for precinct_name, geometry in geometries.items():
        point = geometry.representative_point()
        x, y = projected(point.x, point.y)
        label_draw.text(
            (x, y),
            precinct_name,
            anchor="mm",
            fill=(60, 60, 60, 255),
            font=label_font,
            stroke_width=2,
            stroke_fill=(255, 255, 255, 220),
        )

    legend_width, legend_height = (estimate_svg_legend_box(legend) if legend is not None else (0, 0))
    top_pad = 94
    bottom_pad = max(42, legend_height + 44 if legend is not None else 42)
    side_pad = 28
    canvas_size = (map_image.width + side_pad * 2, map_image.height + top_pad + bottom_pad)
    map_offset = (side_pad, top_pad)

    canvas = Image.new("RGBA", canvas_size, (255, 255, 255, 255))
    canvas.paste(map_image, map_offset, map_image)
    canvas_draw = ImageDraw.Draw(canvas)
    draw_title(canvas_draw, canvas_size, title)
    if legend is not None:
        draw_legend(canvas_draw, canvas_size, legend)

    png_path = output_stem.with_suffix(".png")
    canvas.save(png_path)

    basemap_bytes = io.BytesIO()
    basemap.image.save(basemap_bytes, format="PNG")
    basemap_b64 = base64.b64encode(basemap_bytes.getvalue()).decode("ascii")
    global_projected = lambda lon, lat: (
        map_offset[0] + basemap.project(lon, lat)[0],
        map_offset[1] + basemap.project(lon, lat)[1],
    )

    svg_parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{canvas_size[0]}" height="{canvas_size[1]}" viewBox="0 0 {canvas_size[0]} {canvas_size[1]}">',
        f'<rect width="{canvas_size[0]}" height="{canvas_size[1]}" fill="white"/>',
        f'<text x="{canvas_size[0] / 2:.2f}" y="54" text-anchor="middle" font-family="Arial, sans-serif" font-size="34" font-weight="700" fill="#262626">{title}</text>',
        f'<line x1="30" y1="78" x2="{canvas_size[0] - 30}" y2="78" stroke="#b9b9b9" stroke-width="2"/>',
        f'<image href="data:image/png;base64,{basemap_b64}" x="{map_offset[0]}" y="{map_offset[1]}" width="{basemap.width}" height="{basemap.height}"/>',
    ]

    for precinct_name, geometry in geometries.items():
        fill_rgb = color_fn(value_getter(results[precinct_name]))
        path = geometry_to_svg_path(geometry, global_projected)
        svg_parts.append(
            f'<path d="{path}" fill="rgb{fill_rgb}" fill-opacity="0.75" stroke="#ffffff" stroke-opacity="0.7" stroke-width="2"/>'
        )

    for ward in sorted({result.ward for result in results.values()}):
        ward_geometry = unary_union(
            [geometries[name] for name, result in results.items() if result.ward == ward]
        )
        path = geometry_to_svg_path(ward_geometry, global_projected)
        svg_parts.append(
            f'<path d="{path}" fill="none" stroke="#444444" stroke-width="5"/>'
        )

    city_path = geometry_to_svg_path(city_geometry, global_projected)
    svg_parts.append(f'<path d="{city_path}" fill="none" stroke="#2a2a2a" stroke-width="7"/>')

    for precinct_name, geometry in geometries.items():
        point = geometry.representative_point()
        x, y = global_projected(point.x, point.y)
        svg_parts.append(
            f'<text x="{x:.2f}" y="{y:.2f}" text-anchor="middle" dominant-baseline="middle" '
            'font-family="Arial, sans-serif" font-size="26" font-weight="700" '
            'fill="#3c3c3c" stroke="#ffffff" stroke-width="2" paint-order="stroke fill">'
            f"{precinct_name}</text>"
        )

    if legend is not None:
        box_width, box_height = legend_width, legend_height
        legend_left = canvas_size[0] - box_width - 30
        legend_top = 30 if legend.position != "bottom-right" else canvas_size[1] - box_height - 30
        svg_parts.append(
            f'<rect x="{legend_left}" y="{legend_top}" width="{box_width}" height="{box_height}" rx="16" '
            'fill="white" fill-opacity="0.9" stroke="#5a5a5a" stroke-width="2"/>'
        )
        svg_parts.append(
            f'<text x="{legend_left + 18}" y="{legend_top + 34}" font-family="Arial, sans-serif" '
            'font-size="24" font-weight="700" fill="#2d2d2d">'
            f"{legend.title}</text>"
        )
        for index, (label, color) in enumerate(legend.items):
            y = legend_top + 58 + index * 42
            svg_parts.append(
                f'<rect x="{legend_left + 18}" y="{y}" width="28" height="28" rx="6" '
                f'fill="rgb{color}" stroke="#787878" stroke-width="1"/>'
            )
            svg_parts.append(
                f'<text x="{legend_left + 60}" y="{y + 20}" font-family="Arial, sans-serif" '
                'font-size="19" fill="#373737">'
                f"{label}</text>"
            )

    svg_parts.append("</svg>")
    output_stem.with_suffix(".svg").write_text("\n".join(svg_parts), encoding="utf-8")


def generate_all_outputs() -> None:
    results = load_precinct_results()
    geometries = load_precinct_geometries()
    missing, extra = validate_join(results, geometries)
    if missing or extra:
        raise ValueError(f"join mismatch: missing={missing}, extra={extra}")

    basemap = build_basemap(geometries)
    diff_color_fn, diff_legend = build_difference_colorizer(results)
    render_map(
        "Malden Question 1A yes vote share by precinct",
        lambda result: result.q1a_yes_pct,
        results,
        geometries,
        basemap,
        OUTPUT_DIR / "malden_q1a_precinct_map",
        interpolate_color,
        build_vote_share_legend("Q1A"),
    )
    render_map(
        "Malden Question 1B yes vote share by precinct",
        lambda result: result.q1b_yes_pct,
        results,
        geometries,
        basemap,
        OUTPUT_DIR / "malden_q1b_precinct_map",
        interpolate_color,
        build_vote_share_legend("Q1B"),
    )
    render_map(
        "Malden Question 1A minus 1B yes vote share by precinct",
        lambda result: result.q1a_minus_q1b_yes_pct,
        results,
        geometries,
        basemap,
        OUTPUT_DIR / "malden_q1a_minus_q1b_precinct_map",
        diff_color_fn,
        diff_legend,
    )


if __name__ == "__main__":
    generate_all_outputs()
